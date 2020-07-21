# _*_coding:utf-8 _*_

# @Time:2020/6/27   22:51

# @Author: zlin

#@ File:do_execl.py

#@Software:PyCharm

#@Desc: 读/写execl
import openpyxl
from API_20200720.common import http_request


class Case:
    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None
        self.sql = None

class DoExecl:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.workbook = openpyxl.load_workbook(file_name)
        self.sheet = self.workbook[sheet_name]

    def get_cases(self):
        max_row = self.sheet.max_row
        cases = []
        for r in range(2, max_row+1):
            case = Case()
            case.case_id = self.sheet.cell(row=r, column=1).value
            case.title = self.sheet.cell(row=r, column=2).value
            case.url = self.sheet.cell(row=r, column=3).value
            case.data = self.sheet.cell(row=r, column=4).value
            case.method = self.sheet.cell(row=r, column=5).value
            case.expected = self.sheet.cell(row=r, column=6).value
            case.sql = self.sheet.cell(row=r, column=9).value # sql

            cases.append(case)
        return cases
        self.workbook.close()

    def write_result(self, row, actual, result):
        sheet = self.workbook[self.sheet_name]
        # sheet.cell(row, 6).value = expected
        sheet.cell(row, 7).value = actual
        sheet.cell(row, 8).value = result
        self.workbook.save(filename=self.file_name)
        self.workbook.close()



if __name__ == '__main__':
    from API_20200628.common import contants
    do_execl = DoExecl(contants.case_file, sheet_name='login')
    http_request = http_request.HttpRequest()
    cases = do_execl.get_cases()
    for case in cases:
        # print(case.data)
        # print(case.method)
        print(case.__dict__)
        resp = http_request.request(case.method, case.url, case.data)
        actual = resp.text
        if case.expected == actual:
            do_execl.write_result(case.case_id+1, actual, 'SUCCESS')
        else:
            do_execl.write_result(case.case_id+1, actual, 'FAIL')